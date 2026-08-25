"""
Pillar - Padronizador de Curva ABC (Semana 1 do Método Viga Mestra)
=====================================================================
Converte o relatório "Curva ABC" exportado em PDF pelo sistema
Pontual Tecnologia em uma planilha padronizada (.xlsx), pronta para
alimentar a etapa de análise (giro, margem, produtos isca).

100% determinístico (regex/posição de coluna + pandas) — NENHUMA
chamada de IA é feita aqui. O objetivo é zerar o custo de token
nessa etapa mecânica; a IA entra só na etapa de interpretação/
redação do diagnóstico.

Uso:
    python pillar_padroniza_curva_abc.py entrada.pdf saida.xlsx

O sistema exporta esse relatório em (pelo menos) DOIS layouts
diferentes, dependendo da configuração usada na extração. O script
detecta automaticamente qual é e aplica o parser certo:

  FORMATO SIMPLES  — 2 linhas por produto, cada campo com seu
                     rótulo por extenso ("Unidade UN Qtd. Atual ...").

  FORMATO ESTENDIDO — inclui colunas extras de categoria (Grupo,
                     Sub. Grupo, Linha, Família). Os rótulos não se
                     repetem a cada produto (só uma vez no cabeçalho
                     de cada página), então esse formato é parseado
                     pela POSIÇÃO (x0) de cada palavra na página,
                     reconstruindo as colunas da tabela original.
"""

import re
import sys
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# FORMATO SIMPLES (regex, 2 linhas por produto com rótulos por extenso)
# ---------------------------------------------------------------------------

PRODUTO_RE = re.compile(r"^(\d+)\s+(.+)$")
GRUPO_RE = re.compile(r"^Grupo\.:\s*(\S)")
UNIDADE_RE = re.compile(
    r"^Unidade\s+(\S+)\s+"
    r"Qtd\.\s*Atual\s+([\-\d\.,]+)\s+"
    r"Qtd\.Vend\.\s*([\-\d\.,]*)\s*"
    r"Custo\s+([\-\d\.,]+)\s+"
    r"CustoTotal\s*([\-\d\.,]*)\s*"
    # Em raros casos o extrator de texto do PDF empurra um dígito da
    # frente do número para dentro do rótulo anterior — ex.:
    # "V.l Unit (Méd1)5,77" (valor real: 15,77) ou
    # "Venda Tota1l22.202,00" (valor real: 122.202,00).
    # Os grupos (\d*) abaixo capturam esse dígito perdido para
    # recompor o número certo em vez de descartá-lo.
    r"V\.l Unit \(Méd(\d*)\)\s*([\-\d\.,]+)\s*"
    r"Venda\s*Tota(\d*)l\s*([\-\d\.,]*)\s*"
    r"Margem bruta\s*([\-\d\.,]*)\s*$"
)

COLUNAS_BASE = [
    "codigo", "produto", "grupo_abc", "unidade",
    "qtd_atual", "qtd_vendida", "custo_unit", "custo_total",
    "preco_venda_medio", "venda_total", "margem_bruta_pct",
]
COLUNAS_ESTENDIDAS = COLUNAS_BASE + [
    "codigo_central", "grupo_produto", "sub_grupo", "linha", "familia",
]


def numero_br(valor):
    """Converte string em formato brasileiro ('1.234,56') para float.
    Campo vazio (produto sem movimento) vira None."""
    if not valor:
        return None
    valor = str(valor).replace(".", "").replace(",", ".")
    try:
        return float(valor)
    except ValueError:
        return None


PERIODO_RE = re.compile(
    r"Per[íi]odo\s*(?:de)?:?\s*(\d{2})/(\d{2})/(\d{4})\s*a\s*(\d{2})/(\d{2})/(\d{4})"
)


def extrair_periodo(caminho_pdf: str) -> tuple[str, str] | None:
    """Lê a 1a página em busca de 'Periodo de: DD/MM/AAAA a DD/MM/AAAA'
    (o relatório da Pontual Tecnologia sempre declara o período coberto
    no cabeçalho). Retorna (data_inicio, data_fim) em AAAA-MM-DD, ou
    None se o layout não trouxer essa linha reconhecível."""
    with pdfplumber.open(caminho_pdf) as pdf:
        texto = pdf.pages[0].extract_text() or ""
    m = PERIODO_RE.search(texto)
    if not m:
        return None
    d1, m1, a1, d2, m2, a2 = m.groups()
    return f"{a1}-{m1}-{d1}", f"{a2}-{m2}-{d2}"


TOTAL_GERAL_RE = re.compile(
    r"Total\s*Geral:?\s*([\d\.]+,\d{2})\s+([\d\.]+,\d{2})\s+([\d\.,]+)"
)


def extrair_total_oficial(caminho_pdf: str) -> float | None:
    """Lê a última página em busca de 'Total Geral: <custo> <venda> <margem%>',
    linha de fechamento que o próprio sistema Pontual imprime no relatório.
    Retorna o valor de venda total oficial (float), ou None se não achar.
    Serve para o script conferir sozinho se perdeu produto na extração,
    em vez de confiar cegamente na soma linha a linha."""
    with pdfplumber.open(caminho_pdf) as pdf:
        texto = pdf.pages[-1].extract_text() or ""
    m = TOTAL_GERAL_RE.search(texto)
    if not m:
        return None
    return numero_br(m.group(2))


def extrair_texto_linhas(caminho_pdf: str) -> list[str]:
    """Extrai o texto de cada página do PDF, uma página por vez,
    liberando o cache para não estourar memória em relatórios grandes."""
    linhas = []
    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text() or ""
            linhas.extend(l for l in texto.split("\n") if l.strip())
            pagina.flush_cache()
    return linhas


def detectar_formato(caminho_pdf: str) -> str:
    """'simples' ou 'estendido', olhando o cabeçalho da 1a página."""
    with pdfplumber.open(caminho_pdf) as pdf:
        palavras = pdf.pages[0].extract_words()
    cabecalho = " ".join(w["text"] for w in palavras if w["top"] < 170)
    if "Familia" in cabecalho or "Sub." in cabecalho:
        return "estendido"
    return "simples"


def parsear_formato_simples(linhas: list[str]):
    grupo_atual = None
    registros = []
    falhas = []
    i = 0
    while i < len(linhas):
        linha = linhas[i]

        m_grupo = GRUPO_RE.match(linha)
        if m_grupo:
            grupo_atual = m_grupo.group(1)
            i += 1
            continue

        if linha.startswith("Código"):
            i += 1
            continue

        m_produto = PRODUTO_RE.match(linha)
        if m_produto and i + 1 < len(linhas) and linhas[i + 1].startswith("Unidade"):
            m_dados = UNIDADE_RE.match(linhas[i + 1])
            if m_dados:
                codigo, produto = m_produto.groups()
                (unidade, qtd_atual, qtd_vend, custo, custo_total,
                 digito_perdido_medio, preco_medio,
                 digito_perdido_venda, venda_total,
                 margem) = m_dados.groups()
                preco_medio = (digito_perdido_medio or "") + preco_medio
                if venda_total:
                    venda_total = (digito_perdido_venda or "") + venda_total
                registros.append({
                    "codigo": codigo,
                    "produto": produto.strip(),
                    "grupo_abc": grupo_atual,
                    "unidade": unidade,
                    "qtd_atual": numero_br(qtd_atual),
                    "qtd_vendida": numero_br(qtd_vend),
                    "custo_unit": numero_br(custo),
                    "custo_total": numero_br(custo_total),
                    "preco_venda_medio": numero_br(preco_medio),
                    "venda_total": numero_br(venda_total),
                    "margem_bruta_pct": numero_br(margem),
                })
            else:
                falhas.append((linha, linhas[i + 1]))
            i += 2
        else:
            i += 1

    df = pd.DataFrame(registros, columns=COLUNAS_BASE)
    return df, falhas


# ---------------------------------------------------------------------------
# FORMATO ESTENDIDO (posição de coluna — x0 de cada palavra na página)
# ---------------------------------------------------------------------------

# Limites de coluna calibrados a partir da posição dos rótulos do
# cabeçalho (que se repete, idêntico, em todas as páginas do relatório).
COLS_CATEGORIA = [
    ("codigo", 0, 58),
    ("produto", 143, 400),
    ("grupo_produto", 400, 508),
    ("sub_grupo", 508, 616),
    ("linha", 616, 724),
    ("familia", 724, 10_000),
]
COLS_NUMERICAS = [
    ("qtd_atual", 344, 413),
    ("qtd_vendida", 413, 497),
    ("custo_unit", 497, 551),
    ("custo_total", 551, 616),
    ("preco_venda_medio", 616, 698),
    ("venda_total", 698, 771),
    ("margem_bruta_pct", 771, 10_000),
]
# Unidade normal é só letra (UN, SC, KG, ML...), mas o sistema também usa
# unidade de área/volume com dígito (M2, M3). Sem aceitar o dígito final, a
# linha inteira do produto era descartada (nunca reconhecida como "linha de
# dados"), causando perda silenciosa de produtos inteiros do catálogo,
# sobretudo cerâmica, laje e forro (medidos em m²/m³).
UNIDADE_TOKEN_RE = re.compile(r"^[A-ZÇ]{1,4}[0-9]?$")


def _coluna_de(x0: float, limites) -> str | None:
    for nome, ini, fim in limites:
        if ini <= x0 < fim:
            return nome
    return None


def _agrupar_em_linhas(palavras, tolerancia=3):
    """Agrupa palavras da página em linhas por proximidade vertical (top)."""
    palavras = sorted(palavras, key=lambda w: (w["top"], w["x0"]))
    linhas = []
    linha_atual = []
    top_atual = None
    for w in palavras:
        if top_atual is None or abs(w["top"] - top_atual) <= tolerancia:
            linha_atual.append(w)
            top_atual = w["top"] if top_atual is None else top_atual
        else:
            linhas.append(linha_atual)
            linha_atual = [w]
            top_atual = w["top"]
    if linha_atual:
        linhas.append(linha_atual)
    return linhas


def parsear_formato_estendido(caminho_pdf: str):
    registros = []
    grupo_abc_atual = None
    pendente = None  # produto ainda sem sua linha de números

    def finalizar_pendente_sem_dados():
        # produto sem linha numérica correspondente (não deveria
        # acontecer no layout normal, mas evita perder o registro)
        if pendente is not None:
            registros.append({**pendente, "grupo_abc": grupo_abc_atual})

    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina_idx, pagina in enumerate(pdf.pages):
            # A 1a página tem o cabeçalho completo do relatório (nome da
            # empresa, endereço, "Periodo de:", primeiro "Grupo.:"), que
            # ocupa bem mais espaço vertical que o cabeçalho compacto
            # (só nome das colunas) repetido nas páginas seguintes. Usar
            # o mesmo limite para as duas situações cortava os 1o e 2o
            # produtos de TODA página a partir da 2a (calibrado empiricamente:
            # cabeçalho compacto termina por volta de top=37, 1o produto real
            # de cada página começa em top=70, exatamente onde o limite de
            # 130 usado antes cortava sem necessidade).
            limite_topo_cabecalho = 130 if pagina_idx == 0 else 45
            palavras = pagina.extract_words()
            for linha in _agrupar_em_linhas(palavras):
                linha = sorted(linha, key=lambda w: w["x0"])
                primeira = linha[0]

                # marcador de grupo ABC: "Grupo.: A"
                if primeira["text"] == "Grupo.:":
                    if len(linha) > 1:
                        grupo_abc_atual = linha[1]["text"]
                    continue

                # cabeçalho de página / metadados — ignora
                if primeira["top"] < limite_topo_cabecalho:
                    continue
                if primeira["text"] in (
                    "Código", "Unidade", "Relatório", "Curva",
                    "Página:", "Fone:", "Data:", "Hora:", "Periodo",
                ):
                    continue

                # nova linha de produto: primeiro token é só dígitos
                # e está na coluna do código
                if primeira["x0"] < 58 and re.match(r"^\d+$", primeira["text"]):
                    if pendente is not None:
                        finalizar_pendente_sem_dados()
                    pendente = {
                        "codigo": primeira["text"], "produto": "",
                        "grupo_produto": "", "sub_grupo": "", "linha": "",
                        "familia": "", "codigo_central": "",
                    }
                    for w in linha[1:]:
                        col = _coluna_de(w["x0"], COLS_CATEGORIA)
                        if col and col != "codigo":
                            pendente[col] = (pendente[col] + " " + w["text"]).strip()
                    continue

                # linha de dados numéricos: primeiro token é a unidade
                # (sigla em maiúsculas), na faixa de x0 da coluna Unidade
                if (270 <= primeira["x0"] <= 320
                        and UNIDADE_TOKEN_RE.match(primeira["text"])
                        and pendente is not None):
                    registro = {**pendente, "unidade": primeira["text"],
                                "grupo_abc": grupo_abc_atual}
                    for w in linha[1:]:
                        col = _coluna_de(w["x0"], COLS_NUMERICAS)
                        if col:
                            registro[col] = w["text"]
                    for nome, _, _ in COLS_NUMERICAS:
                        registro[nome] = numero_br(registro.get(nome))
                    registros.append(registro)
                    pendente = None
                    continue

                # continuação (quebra de linha) de um campo de
                # categoria do produto pendente
                if pendente is not None:
                    for w in linha:
                        col = _coluna_de(w["x0"], COLS_CATEGORIA)
                        if col and col != "codigo":
                            pendente[col] = (pendente[col] + " " + w["text"]).strip()

    if pendente is not None:
        finalizar_pendente_sem_dados()

    df = pd.DataFrame(registros, columns=COLUNAS_ESTENDIDAS)
    return df, []


# ---------------------------------------------------------------------------
# Saída padronizada
# ---------------------------------------------------------------------------

def salvar_xlsx(df: pd.DataFrame, caminho_saida: str, periodo: tuple[str, str] | None = None):
    # Linha 1 = banner do período coberto por este relatório (cada PDF
    # vira sua própria planilha, nunca consolidada, porque o cliente
    # não manda cortes de período padronizados). Cabeçalho na linha 2,
    # dados a partir da linha 3.
    df.to_excel(caminho_saida, index=False, sheet_name="Curva ABC padronizada", startrow=1)

    wb = load_workbook(caminho_saida)
    ws = wb.active

    fonte_padrao = Font(name="Arial", size=10)
    fonte_cabecalho = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fonte_periodo = Font(name="Arial", size=10, bold=True)
    preenchimento_cabecalho = PatternFill("solid", fgColor="1F4E78")
    preenchimento_periodo = PatternFill("solid", fgColor="D9E1F2")

    if periodo:
        texto_periodo = f"Período do relatório: {periodo[0]} a {periodo[1]}"
    else:
        texto_periodo = "Período do relatório: não detectado automaticamente no PDF, preencher manualmente"
    ws.cell(row=1, column=1, value=texto_periodo)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    for celula in ws[1]:
        celula.font = fonte_periodo
        celula.fill = preenchimento_periodo

    for celula in ws[2]:
        celula.font = fonte_cabecalho
        celula.fill = preenchimento_cabecalho
        celula.alignment = Alignment(horizontal="center")

    larguras = {
        "codigo": 10, "produto": 45, "grupo_abc": 10, "unidade": 9,
        "qtd_atual": 12, "qtd_vendida": 13, "custo_unit": 12,
        "custo_total": 14, "preco_venda_medio": 16, "venda_total": 14,
        "margem_bruta_pct": 15, "codigo_central": 12, "grupo_produto": 18,
        "sub_grupo": 18, "linha": 16, "familia": 18,
    }
    for idx, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = larguras.get(col, 12)

    for linha in ws.iter_rows(min_row=3):
        for celula in linha:
            celula.font = fonte_padrao

    ws.freeze_panes = "A3"
    wb.save(caminho_saida)


def main():
    if len(sys.argv) < 3:
        print(
            "Uso: python pillar_padroniza_curva_abc.py entrada.pdf saida.xlsx\n"
            "  ou: python pillar_padroniza_curva_abc.py entrada.pdf pasta_destino/\n"
            "      (nesse caso o nome do arquivo é montado sozinho com o período\n"
            "      detectado no PDF, ex: curva-abc-padronizada_2026-04-01_a_2026-06-30.xlsx)"
        )
        sys.exit(1)

    entrada, saida = sys.argv[1], sys.argv[2]
    formato = detectar_formato(entrada)
    print(f"Formato detectado: {formato}")

    periodo = extrair_periodo(entrada)
    if periodo:
        print(f"Período do relatório: {periodo[0]} a {periodo[1]}")
    else:
        print("Período do relatório: não detectado automaticamente, preencher manualmente na planilha")

    if formato == "simples":
        linhas = extrair_texto_linhas(entrada)
        df, falhas = parsear_formato_simples(linhas)
    else:
        df, falhas = parsear_formato_estendido(entrada)

    saida_path = Path(saida)
    if saida_path.suffix.lower() != ".xlsx":
        # 'saida' é uma pasta (cada PDF vira sua própria planilha,
        # nunca consolidada): monta o nome do arquivo com o período.
        if periodo:
            nome = f"curva-abc-padronizada_{periodo[0]}_a_{periodo[1]}.xlsx"
        else:
            nome = "curva-abc-padronizada_periodo-nao-detectado.xlsx"
        saida_path.mkdir(parents=True, exist_ok=True)
        saida = str(saida_path / nome)

    salvar_xlsx(df, saida, periodo)

    print(f"Produtos padronizados: {len(df)}")
    if "grupo_abc" in df.columns:
        print(f"Grupos encontrados: {sorted(df['grupo_abc'].dropna().unique().tolist())}")
    print(f"Linhas com falha de parsing: {len(falhas)}")
    for f in falhas[:10]:
        print("  FALHA:", f)

    total_oficial = extrair_total_oficial(entrada)
    soma_extraida = df["venda_total"].sum() if "venda_total" in df.columns else None
    if total_oficial and soma_extraida is not None:
        cobertura = soma_extraida / total_oficial * 100 if total_oficial else 0
        print(f"Total Geral oficial do relatório (rodapé): R$ {total_oficial:,.2f}")
        print(f"Soma de venda_total extraída: R$ {soma_extraida:,.2f} ({cobertura:.1f}% do oficial)")
        if cobertura < 99.5:
            print(
                f"  ATENÇÃO: cobertura abaixo de 99,5%. Isso indica produto(s) perdido(s) "
                f"na extração (não é diferença de arredondamento). Não usar a soma linha a "
                f"linha como faturamento do período sem investigar antes; use o Total Geral "
                f"oficial (R$ {total_oficial:,.2f}) como fonte do resumo executivo."
            )
    else:
        print("Não foi possível localizar o 'Total Geral' oficial no rodapé para conferência de cobertura.")

    print(f"Arquivo salvo em: {saida}")


if __name__ == "__main__":
    main()
